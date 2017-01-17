#!/usr/bin/env python
# encoding: utf-8
"""
find_largest_positive_change.py

Copyright (c) 2017 Richard Isaacson<ricahrd.c.isaacson@gmail.com>. All rights reserved.
"""

import sys
# Import Library to Support Excel 2010 Workbook
from openpyxl import load_workbook


def load_worksheet(file):
    # Read in workbook and use the default sheet
    wb = load_workbook(file, read_only=True)
    return wb.active


def process_file(ws, row_start, row_end, row_step):
    """
    Process Worksheet from File

    :param ws: Worksheet Object
    :param row_start:
    :param row_end:
    :param row_step:
    :return: largest positive change and the week of the change
    """
    # Initial Values
    largest_positive_change = 0
    largest_positive_change_week = "1970-01-01 00:00:00"
    week_value_previous = 0
    week_value_current = -99999

    # Processing logic.
    # Let's hardcode the range and make assumptions on the data layout
    # as the first pass through it would be difficult to analize the file
    # for data.
    for i in range(row_start, row_end, row_step):
        week_value_previous = week_value_current
        start_column = 'A%s' % i
        end_column = 'E%s' % i
        ((close_date, close, open, high, low),) = ws[start_column:end_column]
        week_value_current = close.value
        if week_value_previous != -99999:
            week_difference_current = week_value_current - week_value_previous
            if week_difference_current > largest_positive_change:
                largest_positive_change = week_difference_current
                largest_positive_change_week = '%s' % close_date.value
    return largest_positive_change, largest_positive_change_week


def main(argv=None):
    """

    :param argv:
    :return: 0 Always
    """

    worksheet = load_worksheet('2016-11_SP_10years1.xlsx')

    largest_positive_change, largest_positive_change_week = process_file(worksheet, 391, 1, -1)

    # Print out what we found to be the largest change and the week of the change.
    print 'largest_positive_change: %s, largest_positive_change_week: %s' \
          % (largest_positive_change, largest_positive_change_week)

    return 0        # success


if __name__ == '__main__':
    status = main()
    sys.exit(status)
